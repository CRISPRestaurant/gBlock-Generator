import os
from math import *
from random import *
import primer3
import openpyxl as pyxl
import pandas as pd
import sys
import matplotlib.pyplot as plt
import time
import seaborn as sns

# libraries cooked up for greater organization and specialized potential in the future
import cooked_up_libraries.dna_properties as dna_properties
import cooked_up_libraries.gene_table as gene_table
import cooked_up_libraries.restriction_enzymes as restriction_enzymes
import cooked_up_libraries.guide_rna_library as guide_rna_library
import cooked_up_libraries.fasta_library as fasta_library
import cooked_up_libraries.bio_string_parser as bsp
import cooked_up_libraries.restriction_digest_analysis as rda

# Ensures all the previous three filters were passed for a given sequence
def firstgBlockFilterOutcome(sequence):
    return dna_properties.noBsaIRestrictionSites(sequence) and dna_properties.noHomopolymerMoreThanFourTs(sequence) and dna_properties.noHomopolymerMoreThanFiveAsAndGs(sequence)

#### Assign Priority System for Guide RNAs ####

# Assigns a higher priority for guides that are earlier
# in the ORF
def obtainEarlierORFPriority(position):
    return (1) / log(position)

# Assigns a higher priority to guides that do not have off-target
# sites
def obtainOffTargetSitePriority(off_target_output_folder, output_number):
    file = open("%s/%s.offtargets" % (off_target_output_folder, output_number), "r")

    first_line = True

    for line in file:
        if first_line:
            first_line = False
        else:
            if "There are no predicted off-targets." in line:
                return 1
            else:
                split_list = line.split(";")
                off_targets = len(split_list)

                return e ** (-1 * off_targets)

# Prioritizes a list of guideRNAs based on the last four
# priority metrics

# sequences is a list in the form of [[first sequence number, first sequence nucleotide ordering, first sequence full_chromosome
# and position, first sequence sense or anti-sense, . . . more metadata . . .]]

# output_folder is where temp files regarding offtarget sites for the guideRNA will find its home

# primer_search_availability_database is a metric to assess how easy it is to find primers that
# can confirm the gene edit took place
def prioritize_guide_RNAs(sequences, off_target_output_folder, primer_database):
    # priority_dict is contained as {guideRNA: priority_value}
    priority_dict = {}
    sequence_properties = {}
    sequenced_list = []

    for sequence in sequences:
        chromosome_and_position = sequence[2].split(":")
        chromosome = chromosome_and_position[0]
        position = int(chromosome_and_position[1])

        gc_priority = dna_properties.obtainGCPriority(sequence[1])
        last_purine_priority = dna_properties.obtainLastPurinePriority(sequence[1])
        earlier_orf_priority = obtainEarlierORFPriority(position)
        off_target_site_priority = obtainOffTargetSitePriority(off_target_output_folder, sequence[0])
        primer_availability_priority = 0

        primer_results = primer_database[sequence[0]]

        for degree_of_asymmetry in primer_results.keys():
            primer_result = primer_results[degree_of_asymmetry]

            for primer_pair in primer_result:
                max_bp_digest_length = max(primer_pair[2][0], primer_pair[2][1])
                min_bp_digest_length = min(primer_pair[2][0], primer_pair[2][1])

                primer_availability_priority += pow(max_bp_digest_length - min_bp_digest_length, 4)

        priority_value = gc_priority * last_purine_priority * earlier_orf_priority * off_target_site_priority * primer_availability_priority

        priority_dict[sequence[1]] = priority_value
        sequence_properties[sequence[1]] = sequence

    # sequenced_list datatype is [(guideRNA nucleotide sequence, priority number)] and is sorted from highest priority
    # to lowest priority
    sequenced_list = sorted(priority_dict.items(), key = lambda x: x[1], reverse = True)

    # We don't need priority number in the output so we're extracting the nucleotide sequence from sequenced_list
    output_sequenced_list = []
    for sequence_and_priority in sequenced_list:
        output_sequenced_list.append(sequence_properties[sequence_and_priority[0]])

    return output_sequenced_list

# END OF METHODS THAT HELP OBTAIN THE PRIORITY OF GUIDERNAS

# Obtain a series of primers from a target strand
def obtainPrimers(target, left_primer_length, right_primer_start, right_primer_length, location):
    raw_results = primer3.bindings.designPrimers(
        {
            'SEQUENCE_ID': 'MH1000',
            'SEQUENCE_TEMPLATE': target,
            'SEQUENCE_PRIMER_PAIR_OK_REGION_LIST': [[0, left_primer_length, right_primer_start, right_primer_length]]
        },
        {
            'PRIMER_OPT_SIZE': 20,
            # 'PRIMER_PICK_INTERNAL_OLIGO': 1,
            # 'PRIMER_INTERNAL_MAX_SELF_END': 8,
            # 'PRIMER_MIN_SIZE': 18,
            # 'PRIMER_MAX_SIZE': 25,
            # 'PRIMER_OPT_TM': 60.0,
            'PRIMER_MIN_TM': 55.0,
            'PRIMER_MAX_TM': 63.0,
            'PRIMER_MIN_GC': 20.0,
            'PRIMER_MAX_GC': 80.0,
            # 'PRIMER_MAX_POLY_X': 100,
            # 'PRIMER_INTERNAL_MAX_POLY_X': 100,
            # 'PRIMER_SALT_MONOVALENT': 50.0,
            # 'PRIMER_DNA_CONC': 50.0,
            # 'PRIMER_MAX_NS_ACCEPTED': 0,
            # 'PRIMER_MAX_SELF_ANY': 12,
            # 'PRIMER_MAX_SELF_END': 8,
            # 'PRIMER_PAIR_MAX_COMPL_ANY': 12,
            # 'PRIMER_PAIR_MAX_COMPL_END': 8,
            # 'PRIMER_PRODUCT_SIZE_RANGE': [[75,100],[100,125],[125,150],
            #                               [150,175],[175,200],[200,225]],
        }
    )

    # Saves explicity primer3 output data to analyze errors
    df = pd.DataFrame(list(raw_results.items()), columns = ["Key", "Value"])
    df.to_csv(location, index = False)

    num_left_primers = raw_results["PRIMER_LEFT_NUM_RETURNED"]
    num_right_primers = raw_results["PRIMER_RIGHT_NUM_RETURNED"]

    min_primers = min(num_left_primers, num_right_primers)

    output = []

    for i in range(min_primers):
        # Obtain left primer sequence
        left_primer = raw_results["PRIMER_LEFT_" + str(i) + "_SEQUENCE"]

        # Obtain left primer melting temperature
        left_primer_tm = raw_results["PRIMER_LEFT_" + str(i) + "_TM"]

        # Obtain left primer position
        left_primer_position_and_length = raw_results["PRIMER_LEFT_" + str(i)]
        left_primer_position = left_primer_position_and_length[0]
        left_primer_length = left_primer_position_and_length[1]

        # Obtain right primer sequence
        right_primer = raw_results["PRIMER_RIGHT_" + str(i) + "_SEQUENCE"]

        # Obtain right primer melting temperature
        right_primer_tm = raw_results["PRIMER_RIGHT_" + str(i) + "_TM"]

        # Obtain right primer position
        right_primer_position_and_length = raw_results["PRIMER_RIGHT_" + str(i)]
        right_primer_position = right_primer_position_and_length[0]
        right_primer_length = right_primer_position_and_length[1]

        output.append([[left_primer, left_primer_tm, left_primer_position, left_primer_length], [right_primer, right_primer_tm, right_primer_position, right_primer_length]])

    return output

# Obtains HI-CRISPR gBlock that does not contain the BsaI endpoints and only contains the
# left homology arm, stop codon, SbfI, right homology arm, and guideRNA
# Also obtains optimal primer information to confirm edit has taken place using SbFI restriction digest and Gel Electrophoresis
def obtaingBlockCandidatesWithoutBsaISite(organism, gene, gene_table_for_organism, guide_rna_results_storage_folder, chopchop_dir, online):
    output = []

    # Very crude processing of primer information
    primer_target_round_one = {}
    # Test info to keep in memory
    gBlocks_without_BsaI_guideRNA_info = {}
    # Very refined version with primer sequence and melting temperature
    primer_target_final_round = {}

    SbfI = "CCTGCAGG"

    guide_rnas = []
    if online:
        guide_rnas = guide_rna_library.obtainGuideRNAsSavedOrOnlineCHOPCHOP(organism, gene, guide_rna_results_storage_folder, time_allotment = 6000)
    else:
        guide_rnas = guide_rna_library.obtainGuideRNAsOfflineCHOPCHOP(organism, gene, guide_rna_results_storage_folder, chopchop_dir)

    organism_fasta = fasta_library.processOrganismFasta(organism)

    first_filter_passed_guideRNAs = []
    guide_rna_to_assembled_sequence_dict = {}

    ## This section processes the guideRNAs and prioritizes
    ## them based on a variety of metrics
    for guide_rna in guide_rnas:
        if guide_rna[3] == "+" and float(guide_rna[10]) > 0:
            full_chromosome = guide_rna[2].split(":")

            chromosome = full_chromosome[0]

            # CHOPCHOP's output position does not follow computer science conventions of 0
            # being the first character in a string
            position = int(full_chromosome[1]) - 1

            left_hr_start_position = position - 50
            left_hr_end_position = position - 1

            left_hr = bsp.substring(organism_fasta[chromosome], position - 1, 50, left_endpoint = False)
            guide_plus_pam = guide_rna[1]
            guide = guide_plus_pam[0: (len(guide_plus_pam) - 3)]
            one_extra = organism_fasta[chromosome][position + len(guide_plus_pam)]

            right_hr_start_position = position + len(guide_plus_pam) + 1
            right_hr_end_position = position + len(guide_plus_pam) + 51

            right_hr = bsp.substring(organism_fasta[chromosome], position + len(guide_plus_pam) + 1, 50, left_endpoint = True)

            exon_ranges = gene_table_for_organism[gene]["exonRanges"]

            # Simple procedure to find exon start position that the guideRNA is residing in
            exon_start_position = 0

            for exon_range in exon_ranges:
                if position >= int(exon_range[0]) and position <= int(exon_range[1]):
                    exon_start_position = int(exon_range[0])
                    break

            difference = position - exon_start_position
            remainder = difference % 3

            prefix = ""

            if remainder != 0:
                prefix = dna_properties.generateRandomSequence(3 - remainder)

            # build variable is easy recognition and analysis by human observers
            build = left_hr + "|" + prefix + "|" + "TAA" + "|" + SbfI + "|" + right_hr + "|||" + guide_rna[1]
            build = "%s|%s|TAA|%s|%s|||%s" % (left_hr, prefix, SbfI, right_hr, guide_rna[1])
            # build real is the real sequence that should be placed into order forms and et cetera
            build_real = build.replace('|', '')

            build_for_edited_chunk = "%s%sTAA%s%s" % (left_hr, prefix, SbfI, right_hr)
            edited_chromosome = organism_fasta[chromosome][0: left_hr_start_position] + build_for_edited_chunk + organism_fasta[chromosome][(right_hr_end_position + 1):]
            base_pair_position_before_digest_cut = left_hr_start_position + len(left_hr) + len(prefix) + 8

            primers_for_each_degree_of_assymetry = {}

            for degree_of_asymmetry in range(2, 6):
                primers_for_each_degree_of_assymetry[degree_of_asymmetry] = rda.obtainPrimersForMeaningfulDigestAnalysis(edited_chromosome, left_hr_start_position, right_hr_end_position, base_pair_position_before_digest_cut, degree_of_asymmetry, minimum_spacing_length, primer_search_space_length, primer_length, maximum_amplicon_length)

            # different_degree_of_assymetry_plots = {"Degree of Assymetry": [], "DNA Digest Base Pair Difference": [], "Primer Pair": []}
            #
            # minimum_spacing_length = 20
            # primer_search_space_length = 200
            # primer_length = 20
            # maximum_amplicon_length = 1000
            # for degree_of_asymmetry in range(2, 6):
            #     primer_output = rda.obtainPrimersForMeaningfulDigestAnalysis(edited_chromosome, left_hr_start_position, right_hr_end_position, base_pair_position_before_digest_cut, degree_of_asymmetry, minimum_spacing_length, primer_search_space_length, primer_length, maximum_amplicon_length)
            #
            #     for primer_index in range(len(primer_output)):
            #         primer = primer_output[primer_index]
            #
            #         degree_of_asymmetry_column = different_degree_of_assymetry_plots["Degree of Assymetry"]
            #         base_pair_difference_column = different_degree_of_assymetry_plots["DNA Digest Base Pair Difference"]
            #         primer_pair_index_column = different_degree_of_assymetry_plots["Primer Pair"]
            #
            #         degree_of_asymmetry_column.append(degree_of_asymmetry)
            #
            #         max_bp_digest_length = max(primer[2][0], primer[2][1])
            #         min_bp_digest_length = min(primer[2][0], primer[2][1])
            #
            #         base_pair_difference_column.append(max_bp_digest_length - min_bp_digest_length)
            #         primer_pair_index_column.append(primer_index + 1)
            #
            #         different_degree_of_assymetry_plots["Degree of Assymetry"] = degree_of_asymmetry_column
            #         different_degree_of_assymetry_plots["DNA Digest Base Pair Difference"] = base_pair_difference_column
            #         different_degree_of_assymetry_plots["Primer Pair"] = primer_pair_index_column
            #
            # degree_of_asymmetry_plot = pd.DataFrame(different_degree_of_assymetry_plots, columns = ["Degree of Assymetry", "DNA Digest Base Pair Difference", "Primer Pair"])
            # sns.set_theme(style = "whitegrid")
            #
            # degree_of_asymmetry_graph = sns.barplot(x = "Degree of Assymetry", y = "DNA Digest Base Pair Difference", hue = "Primer Pair", data = degree_of_asymmetry_plot)
            # degree_of_asymmetry_graph.set_title("")
            # plt.show()
            #
            # end_program = input("End Program?\n")
            #
            # if end_program == "y":
            #     exit()

            if firstgBlockFilterOutcome(build_real):
                first_filter_passed_guideRNAs.append(guide_rna)
                primer_target_round_one[guide_rna[1]] = primers_for_each_degree_of_assymetry
                gBlocks_without_BsaI_guideRNA_info[guide_rna[0]] = build


    ranked_guideRNAs = prioritize_guide_RNAs(first_filter_passed_guideRNAs, "%s/%s/%s/Off_Targets" % (guide_rna_results_storage_folder, organism, gene), primer_guideRNA_optimal_info)
    ## END SECTION

    #print(first_filter_passed_guideRNAs)
    #print(primer_guideRNA_optimal_info)
    #print(gBlocks_without_BsaI_guideRNA_info)
    #print(ranked_guideRNAs)
    # Chooses four highest ranked guideRNAs
    for i in range(0, 4):
        if i < len(ranked_guideRNAs):
            guide_rna = ranked_guideRNAs[i]

            chromosome_and_position = guide_rna[2].split(":")
            chromosome = chromosome_and_position[0]
            position = int(chromosome_and_position[1])

            build = gBlocks_without_BsaI_guideRNA_info[guide_rna[0]]
            output.append(build)
            

    print("Obtained gBlocks for knocking %s in %s" % (gene, organism))
    return [output, primer_target_final_round]

def obtaingBlocks(organism, gene_list, plasmid, gene_table_for_organism, guide_rna_results_storage_folder, chopchop_dir, online):
    # Obtains two sets of four nucleotides needed to glue the ends of the gBlock onto the plasmid
    start_end_BsaI_sites = restriction_enzymes.findBsaISites(plasmid)
    # Pair of start and end nucleotides for BsaI ligation in order of gBlock
    BsaI_four_nucleotides_ordered = []
    # length of nucleotides before and after special 4 nucleotide defining feature
    BsaI_arm_lengths = []
    # List of four nucleotides for BsaI that have already been guessed and should not be generated again for later gBlocks
    BsaI_four_nucleotides_chaos = []

    BsaI_four_nucleotides_chaos.append(start_end_BsaI_sites[0])
    BsaI_four_nucleotides_chaos.append(start_end_BsaI_sites[1])
    if len(gene_list) == 1:
        BsaI_four_nucleotides_ordered = [start_end_BsaI_sites]
        BsaI_arm_lengths = [[24, 26]]
    else:
        for i in range(len(gene_list)):
            if i == 0:
                start_four_nucleotide = start_end_BsaI_sites[0]
                end_four_nucleotide = dna_properties.generateRandomSequenceWithoutRepeatNucleotidesOfMinimalLengthAndWithBlackList(4, 3, BsaI_four_nucleotides_chaos)

                BsaI_four_nucleotides_chaos.append(end_four_nucleotide)

                BsaI_four_nucleotides_ordered.append([start_four_nucleotide, end_four_nucleotide])
                BsaI_arm_lengths.append([24, 29])
            elif i == len(gene_list) - 1:
                start_four_nucleotide = BsaI_four_nucleotides_ordered[-1][1]
                end_four_nucleotide = start_end_BsaI_sites[1]

                BsaI_four_nucleotides_ordered.append([start_four_nucleotide, end_four_nucleotide])
                BsaI_arm_lengths.append([34, 26])
            else:
                start_four_nucleotide = BsaI_four_nucleotides_ordered[-1][1]
                end_four_nucleotide = dna_properties.generateRandomSequenceWithoutRepeatNucleotidesOfMinimalLengthAndWithBlackList(4, 3, BsaI_four_nucleotides_chaos)

                BsaI_four_nucleotides_chaos.append(end_four_nucleotide)

                BsaI_four_nucleotides_ordered.append([start_four_nucleotide, end_four_nucleotide])
                BsaI_arm_lengths.append([34, 26])

    gBlocks = []
    associated_primers = {}

    counter = 1
    for i in range(len(gene_list)):
        gBlocks_raw = obtaingBlockCandidatesWithoutBsaISite(organism, gene_list[i], gene_table_for_organism, guide_rna_results_storage_folder, chopchop_dir, online)
        gBlocks_without_BsaI = gBlocks_raw[0]
        primer_output = gBlocks_raw[1]

        BsaI_nucleotides = BsaI_four_nucleotides_ordered[i]
        BsaI_arm_length = BsaI_arm_lengths[i]
        gBlocks_for_gene = []

        for j in range(len(gBlocks_without_BsaI)):
            build_complete = False
            total_build = ""

            while not build_complete:
                #print("hello!")
                BsaI_left_arm = BsaI_nucleotides[0] + "|" + dna_properties.generateRandomSequenceWithoutRepeatNucleotidesOfMinimalLength(1, 3) + "|GAGACC"
                BsaI_left_arm = dna_properties.generateRandomSequenceWithoutRepeatNucleotidesOfMinimalLength(5, 3) + "|" + BsaI_left_arm
                BsaI_left_arm = BsaI_left_arm + "|" + dna_properties.generateRandomSequenceWithoutRepeatNucleotidesOfMinimalLength(BsaI_arm_length[0] - len("".join(BsaI_left_arm.split("|"))), 3)

                BsaI_right_arm = "GGTCTC|" + dna_properties.generateRandomSequenceWithoutRepeatNucleotidesOfMinimalLength(1, 3) + "|" + BsaI_nucleotides[1]
                BsaI_right_arm = BsaI_right_arm + "|" + dna_properties.generateRandomSequenceWithoutRepeatNucleotidesOfMinimalLength(5, 3)
                BsaI_right_arm = dna_properties.generateRandomSequenceWithoutRepeatNucleotidesOfMinimalLength(BsaI_arm_length[1] - len("".join(BsaI_right_arm.split("|"))), 3) + "|" + BsaI_right_arm

                total_build = BsaI_left_arm + "|" + gBlocks_without_BsaI[j] + "|" + BsaI_right_arm
                total_build_without_pipes = "".join(total_build.split("|"))

                if dna_properties.noHomopolymerMoreThanFourTs(total_build_without_pipes) and dna_properties.noHomopolymerMoreThanFiveAsAndGs(total_build_without_pipes):
                    build_complete = True
                else:
                    #print(total_build)
                    continue

            gBlocks_for_gene.append(total_build)
            associated_primers[total_build] = primer_output[gBlocks_without_BsaI[j]]
            counter = counter + 1


        gBlocks_for_gene.insert(0, gene_list[i])
        gBlocks.append(gBlocks_for_gene)
    return [gBlocks, associated_primers]

start_time = time.time()

organism = "sacCer3"
gene = "YBL011W"
gene_table_for_organism = gene_table.obtainGeneTableInfo(organism)
print(len(gene_table_for_organism.keys()))
plasmid = "pCRCT"

gene_list = list(gene_table_for_organism.keys())[0:100]
gBlocks_and_primers = obtaingBlocks(organism, gene_list, plasmid, gene_table_for_organism, "GuideRNA_Archives", "chopchop", False)
file_name = "The Mega 100 Gene Knockout"

file = open("Test Output Garbage Dump/HI-CRISPR Knockout/" + file_name + ".txt", "w")

primer_wb = pyxl.Workbook()
ws = primer_wb[primer_wb.sheetnames[0]]

# Assigns A1 to "Name"
wcell_name = ws.cell(1, 1)
wcell_name.value = "Name"

# Assigns B1 to "Sequence"
wcell_sequence = ws.cell(1, 2)
wcell_sequence.value = "Sequence"

# Assigns C1 to "Scale"
wcell_scale = ws.cell(1, 3)
wcell_scale.value = "Scale"

# Assigns D1 to "Purification"
wcell_purification = ws.cell(1, 4)
wcell_purification.value = "Purification"

gene_primers_added_to_excel = []
# Outputs all the gBlock generation and primer verifiers into different
# text files and excel documents
#
# Look at Test Output Garbage Dump/HI-CRISPR Knockout for gBlocks and corresponding primers and amplicons.
for gene_position in range(len(gBlocks_and_primers[0])):
    gene = gBlocks_and_primers[0][gene_position]
    file.write(gene[0] + "\n\n")
    for i in range(1, len(gene)):
        file.write("gBlock #" + str(i) + "\n\n")
        file.write(gene[i] + "\n\n")

        primer_information = gBlocks_and_primers[1][gene[i]]
        file.write(primer_information[0] + "\n\n")

        primer_list = primer_information[1]
        SbFI_digest_calculations = primer_information[2]
        if len(primer_list) == 0:
            print("%s has no primers and sucks butt!" % (gene))
        for j in range(len(primer_list)):
            file.write("Left Primer #" + str(j + 1) + ": " + primer_list[j][0][0] + "\n")
            file.write("Left Primer #" + str(j + 1) + " Melting Temperature: " + str(primer_list[j][0][1]) + "\n\n")

            file.write("Right Primer #" + str(j + 1) + ": " + primer_list[j][1][0] + " (search for " + dna_properties.getAntiSenseStrand(primer_list[j][1][0]) + ")" + "\n")
            file.write("Right Primer #" + str(j + 1) + ": " + str(primer_list[j][1][1]) + "\n\n")

            file.write("5' Sense Restriction Digest End Length: %s\n" % (SbFI_digest_calculations[j][0]))
            file.write("3' Sense Restriction Digest End Length: %s\n" % (SbFI_digest_calculations[j][1]))
            file.write("log(larger DNA digest bp) / log(smaller DNA digest bp): %s\n\n" % (SbFI_digest_calculations[j][2]))

            if j == 0 and not gene[0] in gene_primers_added_to_excel:
                left_primer_row_num = 2 * gene_position + 2
                right_primer_row_num = 2 * gene_position + 3

                # Assigns left primer row values

                wcell_lp_name = ws.cell(left_primer_row_num, 1)
                wcell_lp_name.value = gene[0] + "_LEFT_PRIMER"

                wcell_lp_sequence = ws.cell(left_primer_row_num, 2)
                wcell_lp_sequence.value = primer_list[j][0][0]

                wcell_lp_scale = ws.cell(left_primer_row_num, 3)
                wcell_lp_scale.value = "25nm"

                wcell_lp_purification = ws.cell(left_primer_row_num, 4)
                wcell_lp_purification.value = "STD"

                # Assigns right primer row values

                wcell_rp_name = ws.cell(right_primer_row_num, 1)
                wcell_rp_name.value = gene[0] + "_RIGHT_PRIMER"

                wcell_rp_sequence = ws.cell(right_primer_row_num, 2)
                wcell_rp_sequence.value = primer_list[j][1][0]

                wcell_rp_scale = ws.cell(right_primer_row_num, 3)
                wcell_rp_scale.value = "25nm"

                wcell_rp_purification = ws.cell(right_primer_row_num, 4)
                wcell_rp_purification.value = "STD"

                gene_primers_added_to_excel.append(gene[0])

    file.write("\n\n")

primer_wb.save("Test Output Garbage Dump/HI-CRISPR Knockout/" + file_name + ".xlsx")
file.close()

end_time = time.time()

print((end_time - start_time) / 60)
